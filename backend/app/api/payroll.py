from datetime import date, datetime

import csv
import io
import re
import pyzipper
from fastapi.responses import Response

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.core import User, UserRole

router = APIRouter()


def _roles(db: Session, user: User) -> set[str]:
    return {ur.role.name for ur in db.query(UserRole).filter(UserRole.user_id == user.id).all()}


def _is_finance_employee(db: Session, user_id: int) -> bool:
    row = db.execute(text("""
        SELECT employee_role FROM employee_users
        WHERE user_id = :uid AND COALESCE(is_active, TRUE) = TRUE
        LIMIT 1
    """), {"uid": user_id}).mappings().first()
    return bool(row and "finance" in str(row["employee_role"] or "").strip().lower())


def _franchise_id_for_user(db: Session, user: User):
    r = _roles(db, user)
    if "FranchiseUser" in r:
        row = db.execute(text("SELECT id FROM franchise_users WHERE user_id = :uid AND COALESCE(is_active, TRUE)=TRUE LIMIT 1"), {"uid": user.id}).mappings().first()
        return row["id"] if row else None
    if "ManagerUser" in r or "EmployeeUser" in r:
        row = db.execute(text("""
            SELECT franchise_user_id FROM manager_users WHERE user_id = :uid AND COALESCE(is_active, TRUE)=TRUE
            UNION ALL
            SELECT franchise_user_id FROM employee_users WHERE user_id = :uid AND COALESCE(is_active, TRUE)=TRUE
            LIMIT 1
        """), {"uid": user.id}).mappings().first()
        return row["franchise_user_id"] if row else None
    return None


def _can_payroll(db: Session, user: User) -> bool:
    r = _roles(db, user)
    return "SuperUser" in r or "FranchiseUser" in r or _is_finance_employee(db, user.id)


def _is_superuser(db: Session, user: User) -> bool:
    return "SuperUser" in _roles(db, user)


def _is_franchise_user(db: Session, user: User) -> bool:
    return "FranchiseUser" in _roles(db, user)


def _is_finance_user(db: Session, user: User) -> bool:
    return _is_finance_employee(db, user.id)


def _ensure_tables(db: Session):
    
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS payroll_imports (
            id SERIAL PRIMARY KEY,
            filename VARCHAR(255) NOT NULL,
            payroll_month DATE NULL,
            franchise_user_id INTEGER NULL,
            imported_by_user_id INTEGER NOT NULL,
            imported_at TIMESTAMP NOT NULL DEFAULT NOW(),
            rows_total INTEGER NOT NULL DEFAULT 0,
            rows_matched INTEGER NOT NULL DEFAULT 0,
            status VARCHAR(40) NOT NULL DEFAULT 'processed'
        )
    """))
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS payroll_import_rows (
            id SERIAL PRIMARY KEY,
            import_id INTEGER NOT NULL REFERENCES payroll_imports(id) ON DELETE CASCADE,
            row_number INTEGER NOT NULL,
            matched_user_id INTEGER NULL,
            employee_key VARCHAR(255) NULL,
            employee_name VARCHAR(255) NULL,
            email VARCHAR(255) NULL,
            match_method VARCHAR(80) NULL,
            status VARCHAR(40) NOT NULL DEFAULT 'unmatched',
            message TEXT NULL,
            created_at TIMESTAMP NOT NULL DEFAULT NOW()
        )
    """))

    db.execute(text("""
        CREATE TABLE IF NOT EXISTS payroll_payslips (
            id SERIAL PRIMARY KEY,
            import_id INTEGER NULL REFERENCES payroll_imports(id) ON DELETE SET NULL,
            user_id INTEGER NOT NULL,
            franchise_user_id INTEGER NULL,
            employee_key VARCHAR(255) NULL,
            original_filename VARCHAR(255) NOT NULL,
            zip_filename VARCHAR(255) NULL,
            file_content BYTEA NOT NULL,
            content_type VARCHAR(120) NOT NULL DEFAULT 'application/zip',
            uploaded_by_user_id INTEGER NOT NULL,
            uploaded_at TIMESTAMP NOT NULL DEFAULT NOW()
        )
    """))

    
    for stmt in [
        
        "ALTER TABLE payroll_imports ADD COLUMN IF NOT EXISTS rows_total INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE payroll_imports ADD COLUMN IF NOT EXISTS rows_matched INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE payroll_imports ADD COLUMN IF NOT EXISTS status VARCHAR(40) NOT NULL DEFAULT 'processed'",
        "ALTER TABLE payroll_imports ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP NULL",
        "ALTER TABLE payroll_imports ADD COLUMN IF NOT EXISTS deleted_at TIMESTAMP NULL",
        "ALTER TABLE payroll_imports ADD COLUMN IF NOT EXISTS deleted_by_user_id INTEGER NULL",
        "ALTER TABLE payroll_import_rows ADD COLUMN IF NOT EXISTS matched_user_id INTEGER NULL",
        "ALTER TABLE payroll_import_rows ADD COLUMN IF NOT EXISTS employee_name VARCHAR(255) NULL",
        "ALTER TABLE payroll_import_rows ADD COLUMN IF NOT EXISTS status VARCHAR(40) NOT NULL DEFAULT 'unmatched'",
        "ALTER TABLE payroll_import_rows ADD COLUMN IF NOT EXISTS message TEXT NULL",
        "ALTER TABLE payroll_import_rows ADD COLUMN IF NOT EXISTS match_method VARCHAR(80) NULL",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS import_id INTEGER NULL",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS user_id INTEGER NULL",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS franchise_user_id INTEGER NULL",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS employee_key VARCHAR(255) NULL",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS original_filename VARCHAR(255) NULL",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS zip_filename VARCHAR(255) NULL",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS file_content BYTEA NULL",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS content_type VARCHAR(120) NOT NULL DEFAULT 'application/zip'",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS uploaded_by_user_id INTEGER NULL",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS uploaded_at TIMESTAMP NOT NULL DEFAULT NOW()",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS is_active BOOLEAN NOT NULL DEFAULT TRUE",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS deleted_at TIMESTAMP NULL",
        "ALTER TABLE payroll_payslips ADD COLUMN IF NOT EXISTS deleted_by_user_id INTEGER NULL",
    ]:
        db.execute(text(stmt))
    db.commit()


def _staff_scope_sql(db: Session, current_user: User):
    roles = _roles(db, current_user)
    if "SuperUser" in roles:
        return "1=1", {}
    # Finance users handle payroll documents and must be able to allocate imports
    # to every manager/employee record. Franchise users remain scoped to their
    # own franchise staff.
    if _is_finance_employee(db, current_user.id):
        return "1=1", {}
    fid = _franchise_id_for_user(db, current_user)
    if not fid:
        return "1=0", {}
    return "staff.franchise_user_id = :fid", {"fid": fid}


def _staff_rows(db: Session, current_user: User):
    where, params = _staff_scope_sql(db, current_user)
    rows = db.execute(text(f"""
        WITH staff AS (
            SELECT u.id AS user_id, e.franchise_user_id, e.name, e.surname, e.employee_role AS role, e.email,
                   e.employee_number, 'Employee' AS staff_type
            FROM employee_users e JOIN users u ON u.id = e.user_id
            WHERE COALESCE(e.is_active, TRUE) = TRUE AND COALESCE(u.is_active, TRUE) = TRUE
            UNION ALL
            SELECT u.id AS user_id, m.franchise_user_id, m.name, m.surname, 'Manager' AS role, m.email,
                   m.employee_number AS employee_number, 'Manager' AS staff_type
            FROM manager_users m JOIN users u ON u.id = m.user_id
            WHERE COALESCE(m.is_active, TRUE) = TRUE AND COALESCE(u.is_active, TRUE) = TRUE
        )
        SELECT staff.*
        FROM staff
        WHERE {where}
        ORDER BY staff.name, staff.surname
    """), params).mappings().all()
    return [dict(r) for r in rows]


def _normalise_header(value: str) -> str:
    return re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower()).strip('_')


def _to_float(value):
    if value is None:
        return None
    text_value = str(value).strip()
    if text_value == '':
        return None
    text_value = text_value.replace('R', '').replace('r', '').replace(',', '').replace('%', '').strip()
    try:
        return float(text_value)
    except ValueError:
        return None


def _clean_match_text(value):
    return re.sub(r'\s+', ' ', str(value or '').strip().lower())


def _first(row: dict, *names):
    for name in names:
        if name in row and str(row[name]).strip() != '':
            return row[name]
    return ''


def _read_payroll_file(file: UploadFile):
    content = file.file.read()
    filename = (file.filename or '').lower()
    if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise HTTPException(status_code=400, detail='Excel import needs openpyxl installed. Add openpyxl to requirements.txt and reinstall backend requirements.') from exc
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_normalise_header(h) for h in rows[0]]
        parsed = []
        for idx, values in enumerate(rows[1:], start=2):
            if not values or not any(v is not None and str(v).strip() for v in values):
                continue
            parsed.append({'_row_number': idx, **{headers[i]: values[i] if i < len(values) else '' for i in range(len(headers))}})
        return parsed
    if filename.endswith('.xls'):
        raise HTTPException(status_code=400, detail='Old .xls files are not supported directly. Save the payroll document as .xlsx or CSV, then import again.')
    
    if filename.endswith('.zip'):
        extracted_rows = []
        outer_name = (file.filename or '').rsplit('/', 1)[-1].rsplit('\\', 1)[-1]
        outer_stem = re.sub(r'\.zip$', '', outer_name, flags=re.IGNORECASE)

        with pyzipper.AESZipFile(io.BytesIO(content)) as z:
            pdf_files = [f for f in z.namelist() if f.lower().endswith('.pdf')]
            if not pdf_files:
                pdf_files = [outer_name]
            for idx, pdf_name in enumerate(pdf_files, start=1):
                base = pdf_name.split('/')[-1].replace('.PDF', '').replace('.pdf', '')
                tokens = []
                for c in [base, outer_stem]:
                    c = str(c or '').strip()
                    tokens.append(c)
                    tokens.extend(re.split(r'[^A-Za-z0-9]+', c))
                    if '-' in c:
                        tokens.extend([c.split('-')[0], c.split('-')[-1]])
                employee_code = next((t for t in tokens if t and re.search(r'[A-Za-z]', t) and re.search(r'\d', t)), '')
                if not employee_code:
                    employee_code = next((t for t in tokens if t and re.search(r'\d', t)), outer_stem)
                extracted_rows.append({
                    '_row_number': idx,
                    'empl_no': employee_code,
                    'employee_number': employee_code,
                    'employee_id': employee_code,
                    'payslip_filename': pdf_name if pdf_name.lower().endswith('.pdf') else outer_name,
                    'payslip_zip_bytes': content,
                })
        return extracted_rows


    if filename.endswith('.pdf'):
        raise HTTPException(
            status_code=400,
            detail='Please upload the ZIP payroll export instead of the raw PDF.'
        )
    text_content = content.decode('utf-8-sig', errors='replace')
    sample = text_content[:2048]
    delimiter = ','
    if sample.count(';') > sample.count(','):
        delimiter = ';'
    if sample.count('\t') > sample.count(delimiter):
        delimiter = '\t'
    reader = csv.DictReader(io.StringIO(text_content), delimiter=delimiter)
    parsed = []
    for idx, row in enumerate(reader, start=2):
        norm = {_normalise_header(k): v for k, v in row.items() if k is not None}
        if any(str(v or '').strip() for v in norm.values()):
            norm['_row_number'] = idx
            parsed.append(norm)
    return parsed


def _last_7_digits(value) -> str:
    digits = re.sub(r'\D+', '', str(value or ''))
    return digits[-7:] if len(digits) >= 7 else ''


def _code_variants(value) -> set[str]:
    raw = str(value or '').strip().upper()
    compact = re.sub(r'[^A-Z0-9]+', '', raw)
    digits = re.sub(r'\D+', '', raw)
    variants = {raw, compact, raw[-6:], raw[-7:], compact[-6:], compact[-7:]}
    variants.update(t for t in re.split(r'[^A-Z0-9]+', raw) if t)
    if '-' in raw:
        variants.add(raw.split('-')[0])
        variants.add(raw.split('-')[-1])
    if digits:
        variants.update({digits, digits[-6:], digits[-7:]})
    return {v for v in variants if v}


def _staff_match_maps(staff_rows):
    by_id, by_email, by_name = {}, {}, {}
    by_employee_last7 = {}
    by_employee_code = {}

    for s in staff_rows:
        uid = int(s['user_id'])
        by_id[str(uid)] = s

        email = str(s.get('email') or '').strip().lower()
        if email:
            by_email[email] = s

        full_name = _clean_match_text(
            ' '.join(x for x in [s.get('name'), s.get('surname')] if x)
        )
        if full_name:
            by_name[full_name] = s

        employee_number = str(s.get('employee_number') or '').strip().upper()

        if employee_number:
            for variant in _code_variants(employee_number):
                by_employee_code[variant] = s

        employee_last7 = _last_7_digits(employee_number)
        if employee_last7:
            by_employee_last7[employee_last7] = s

    return by_id, by_email, by_name, by_employee_last7, by_employee_code


def _match_payroll_staff(
    row: dict,
    by_id: dict,
    by_email: dict,
    by_name: dict,
    by_employee_last7: dict,
    by_employee_code: dict,
):
    user_key = str(_first(
        row,
        'empl_no',
        'empl_no_',
        'employee_no',
        'employee_number',
        'employee_id',
        'staff_id',
        'emp_id',
        'personnel_number',
        'user_id'
    )).strip()

    email = str(_first(
        row,
        'email',
        'email_address',
        'employee_email',
        'work_email'
    )).strip().lower()

    employee_name = str(_first(
        row,
        'employee',
        'employee_name',
        'staff_name',
        'name',
        'full_name',
        'employee_full_name'
    )).strip()

    user_key_variants = _code_variants(user_key)

    for variant in user_key_variants:
        if variant and variant in by_employee_code:
            return by_employee_code[variant], 'employee_code', user_key, email, employee_name

    user_key_last7 = _last_7_digits(user_key)

    if user_key_last7 and user_key_last7 in by_employee_last7:
        return by_employee_last7[user_key_last7], 'employee_number_last_7', user_key, email, employee_name

    if user_key and user_key in by_id:
        return by_id[user_key], 'user_id', user_key, email, employee_name

    if email and email in by_email:
        return by_email[email], 'email', user_key, email, employee_name

    cleaned_name = _clean_match_text(employee_name)

    if cleaned_name and cleaned_name in by_name:
        return by_name[cleaned_name], 'full_name', user_key, email, employee_name

    return None, 'unmatched', user_key, email, employee_name


@router.post('/import-document')
def import_payroll_document(
    file: UploadFile = File(...),
    payroll_month: str = Form(''),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _can_payroll(db, current_user):
        raise HTTPException(status_code=403, detail='Payslip import access denied')

    _ensure_tables(db)
    
    rows = _read_payroll_file(file)
    staff_rows = _staff_rows(db, current_user)
    by_id, by_email, by_name, by_employee_last7, by_employee_code = _staff_match_maps(staff_rows)
    current_fid = _franchise_id_for_user(db, current_user)
    imported_at = datetime.utcnow()
    month_date = None
    if payroll_month:
        try:
            parsed_month = date.fromisoformat(payroll_month[:10])
            month_date = date(parsed_month.year, parsed_month.month, 1)
        except ValueError:
            month_date = None
    import_row = db.execute(text("""
        INSERT INTO payroll_imports (filename, payroll_month, franchise_user_id, imported_by_user_id, imported_at, rows_total, rows_matched)
        VALUES (:filename, :payroll_month, :franchise_user_id, :imported_by_user_id, :imported_at, 0, 0)
        RETURNING id
    """), {
        'filename': file.filename or 'payroll-import',
        'payroll_month': month_date,
        'franchise_user_id': current_fid,
        'imported_by_user_id': current_user.id,
        'imported_at': imported_at,
    }).mappings().first()
    import_id = import_row['id']
    results = []
    matched_count = 0
    for row in rows:
        row_no = int(row.get('_row_number') or len(results) + 1)
        staff, match_method, user_key, email, employee_name = _match_payroll_staff(
        row,
        by_id,
        by_email,
        by_name,
        by_employee_last7,
        by_employee_code,
    )
        matched_user_id = int(staff['user_id']) if staff else None
        status = 'matched' if staff else 'unmatched'
        msg = f'Payslip matched by {match_method}' if staff else 'No matching employee in your allowed franchise scope'

        if staff:
            payslip_bytes = row.get('payslip_zip_bytes')
            payslip_filename = row.get('payslip_filename')
            
            if payslip_bytes and payslip_filename:
                db.execute(text("""
                    INSERT INTO payroll_payslips (
                        import_id, user_id, franchise_user_id, employee_key,
                        original_filename, zip_filename, file_content,
                        content_type, uploaded_by_user_id, uploaded_at
                    )
                    VALUES (
                        :import_id, :user_id, :franchise_user_id, :employee_key,
                        :original_filename, :zip_filename, :file_content,
                        :content_type, :uploaded_by_user_id, :uploaded_at
                    )
                """), {
                    'import_id': import_id,
                    'user_id': matched_user_id,
                    'franchise_user_id': staff.get('franchise_user_id'),
                    'employee_key': user_key,
                    'original_filename': payslip_filename,
                    'zip_filename': file.filename or 'payroll.zip',
                    'file_content': payslip_bytes,
                    'content_type': 'application/zip',
                    'uploaded_by_user_id': current_user.id,
                    'uploaded_at': imported_at,
                })
            matched_count += 1
            
        db.execute(text("""
            INSERT INTO payroll_import_rows (
                import_id,
                row_number,
                matched_user_id,
                employee_key,
                employee_name,
                email,
                match_method,
                status,
                message,
                created_at
            )
            VALUES (
                :import_id,
                :row_number,
                :matched_user_id,
                :employee_key,
                :employee_name,
                :email,
                :match_method,
                :status,
                :message,
                :created_at
            )
        """), {
            'import_id': import_id,
            'row_number': row_no,
            'matched_user_id': matched_user_id,
            'employee_key': user_key or None,
            'employee_name': employee_name or None,
            'email': email or None,
            'match_method': match_method,
            'status': status,
            'message': msg,
            'created_at': imported_at,
        })
        
        results.append({
            'row_number': row_no,
            'status': status,
            'message': msg,
            'matched_user_id': matched_user_id,
            'employee_name': employee_name or ((' '.join(x for x in [staff.get('name'), staff.get('surname')] if x)) if staff else ''),
            'email': email,
            'match_method': match_method,
        })

        
    db.execute(text("UPDATE payroll_imports SET rows_total = :total, rows_matched = :matched WHERE id = :id"), {
        'total': len(results), 'matched': matched_count, 'id': import_id
    })
    db.commit()
    return {'import_id': import_id, 'filename': file.filename, 'rows_total': len(results), 'rows_matched': matched_count, 'rows': results}


@router.get('/imports')
def payroll_imports(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_tables(db)

    if _can_payroll(db, current_user):
        where, params = _payslip_management_filter(db, current_user, 'p')
        rows = db.execute(text(f"""
            SELECT DISTINCT i.*
            FROM payroll_imports i
            LEFT JOIN payroll_payslips p ON p.import_id = i.id AND COALESCE(p.is_active, TRUE) = TRUE
            WHERE i.deleted_at IS NULL AND ({where})
            ORDER BY i.imported_at DESC
            LIMIT 100
        """), params).mappings().all()
    else:
        rows = db.execute(text("""
            SELECT DISTINCT i.*
            FROM payroll_imports i
            JOIN payroll_import_rows r ON r.import_id = i.id
            WHERE i.deleted_at IS NULL AND r.matched_user_id = :uid
            ORDER BY i.imported_at DESC
            LIMIT 100
        """), {'uid': current_user.id}).mappings().all()
    return [dict(r) for r in rows]


def _get_visible_import(db: Session, current_user: User, import_id: int):
    if _can_payroll(db, current_user):
        where, params = _payslip_management_filter(db, current_user, 'p')
        params['id'] = import_id
        row = db.execute(text(f"""
            SELECT DISTINCT i.* FROM payroll_imports i
            LEFT JOIN payroll_payslips p ON p.import_id = i.id AND COALESCE(p.is_active, TRUE) = TRUE
            WHERE i.id = :id AND i.deleted_at IS NULL AND ({where})
            LIMIT 1
        """), params).mappings().first()
    else:
        row = db.execute(text("""
            SELECT DISTINCT i.* FROM payroll_imports i
            JOIN payroll_import_rows r ON r.import_id = i.id
            WHERE i.id = :id AND i.deleted_at IS NULL AND r.matched_user_id = :uid
            LIMIT 1
        """), {'id': import_id, 'uid': current_user.id}).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail='Payroll import not found')
    return row


def _get_manageable_import(db: Session, current_user: User, import_id: int):
    if not _can_payroll(db, current_user):
        raise HTTPException(status_code=403, detail='Only SuperUser, FranchiseUser or Finance users can edit/delete payroll imports')
    return _get_visible_import(db, current_user, import_id)


@router.get('/imports/{import_id}')
def payroll_import_detail(import_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_tables(db)
    imp = dict(_get_visible_import(db, current_user, import_id))
    if _can_payroll(db, current_user):
        row_where = 'import_id = :id'
        row_params = {'id': import_id}
    else:
        row_where = 'import_id = :id AND matched_user_id = :uid'
        row_params = {'id': import_id, 'uid': current_user.id}
    rows = db.execute(text(f"""
        SELECT id, row_number, matched_user_id, employee_key, employee_name, email,
               match_method, status, message, created_at
        FROM payroll_import_rows
        WHERE {row_where}
        ORDER BY row_number ASC, id ASC
    """), row_params).mappings().all()
    imp['rows'] = [dict(r) for r in rows]
    return imp


@router.put('/imports/{import_id}')
def update_payroll_import(import_id: int, payload: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_tables(db)
    _get_manageable_import(db, current_user, import_id)
    payroll_month = payload.get('payroll_month') or None
    status = str(payload.get('status') or 'processed')[:40]
    month_date = None
    if payroll_month:
        try:
            parsed_month = date.fromisoformat(str(payroll_month)[:10])
            month_date = date(parsed_month.year, parsed_month.month, 1)
        except ValueError:
            raise HTTPException(status_code=400, detail='Invalid payroll month')
    db.execute(text("""
        UPDATE payroll_imports
        SET payroll_month = :payroll_month, status = :status, updated_at = :updated_at
        WHERE id = :id
    """), {'id': import_id, 'payroll_month': month_date, 'status': status, 'updated_at': datetime.utcnow()})
    db.commit()
    return {'message': 'Payroll import updated'}


@router.delete('/imports/{import_id}')
def delete_payroll_import(import_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_tables(db)
    _get_manageable_import(db, current_user, import_id)
    now = datetime.utcnow()
    db.execute(text("""
        UPDATE payroll_payslips
        SET is_active = FALSE, deleted_at = :deleted_at, deleted_by_user_id = :deleted_by
        WHERE import_id = :id
    """), {'id': import_id, 'deleted_at': now, 'deleted_by': current_user.id})
    db.execute(text("""
        UPDATE payroll_imports
        SET deleted_at = :deleted_at, deleted_by_user_id = :deleted_by, status = 'deleted'
        WHERE id = :id
    """), {'id': import_id, 'deleted_at': now, 'deleted_by': current_user.id})
    db.commit()
    return {'message': 'Payroll import and linked payslips deleted'}


def _payslip_management_filter(db: Session, current_user: User, alias: str = 'p'):
    roles = _roles(db, current_user)
    if 'SuperUser' in roles or _is_finance_employee(db, current_user.id):
        return '1=1', {}
    fid = _franchise_id_for_user(db, current_user)
    if 'FranchiseUser' in roles and fid:
        return f'{alias}.franchise_user_id = :fid', {'fid': fid}
    return '1=0', {}


def _payslip_scope_filter(db: Session, current_user: User, alias: str = 'p'):
    # Managers and employees can ONLY see their own payslip documents.
    # Admin/franchise/finance users may manage linked manager/employee documents
    # through the same endpoints, scoped by role.
    if _can_payroll(db, current_user):
        return _payslip_management_filter(db, current_user, alias)
    return f'{alias}.user_id = :uid', {'uid': current_user.id}


def _get_manageable_payslip(db: Session, current_user: User, payslip_id: int):
    where, params = _payslip_management_filter(db, current_user, 'p')
    params['id'] = payslip_id
    row = db.execute(text(f"""
        SELECT p.id
        FROM payroll_payslips p
        WHERE p.id = :id
          AND COALESCE(p.is_active, TRUE) = TRUE
          AND {where}
        LIMIT 1
    """), params).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail='Payslip not found')
    return row

def _get_scoped_payslip(db: Session, current_user: User, payslip_id: int):
    where, params = _payslip_scope_filter(db, current_user, 'p')
    params['id'] = payslip_id
    row = db.execute(text(f"""
        SELECT p.*
        FROM payroll_payslips p
        WHERE p.id = :id
          AND COALESCE(p.is_active, TRUE) = TRUE
          AND {where}
        LIMIT 1
    """), params).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail='Payslip not found')
    return row

def _is_staff_user(db: Session, user_id: int) -> bool:
    row = db.execute(text("""
        SELECT 1 FROM employee_users WHERE user_id = :uid AND COALESCE(is_active, TRUE)=TRUE
        UNION ALL
        SELECT 1 FROM manager_users WHERE user_id = :uid AND COALESCE(is_active, TRUE)=TRUE
        LIMIT 1
    """), {"uid": user_id}).first()
    return bool(row)


@router.get('/payslips')
def scoped_payslips(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_tables(db)
    if not _is_staff_user(db, current_user.id) and not _can_payroll(db, current_user):
        raise HTTPException(status_code=403, detail='Payslip access denied')
    where, params = _payslip_scope_filter(db, current_user, 'p')
    rows = db.execute(text(f"""
        SELECT p.id, p.employee_key, p.original_filename, p.zip_filename, p.uploaded_at,
               p.user_id, p.franchise_user_id,
               COALESCE(e.name || ' ' || e.surname, m.name || ' ' || m.surname, u.full_name, u.email) AS staff_name
        FROM payroll_payslips p
        LEFT JOIN users u ON u.id = p.user_id
        LEFT JOIN employee_users e ON e.user_id = p.user_id
        LEFT JOIN manager_users m ON m.user_id = p.user_id
        WHERE COALESCE(p.is_active, TRUE) = TRUE AND {where}
        ORDER BY p.uploaded_at DESC
        LIMIT 500
    """), params).mappings().all()
    return [dict(r) for r in rows]


@router.get('/my-payslips')
def my_payslips(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    if not _is_staff_user(db, current_user.id) and not _can_payroll(db, current_user):
        raise HTTPException(status_code=403, detail='Payslip access denied')

    rows = db.execute(text("""
        SELECT
            id,
            employee_key,
            original_filename,
            zip_filename,
            uploaded_at
        FROM payroll_payslips
        WHERE user_id = :uid AND COALESCE(is_active, TRUE) = TRUE
        ORDER BY uploaded_at DESC
    """), {
        "uid": current_user.id
    }).mappings().all()

    return [dict(r) for r in rows]


@router.get('/payslips/{payslip_id}')

def download_payslip(
    payslip_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    if not _is_staff_user(db, current_user.id) and not _can_payroll(db, current_user):
        raise HTTPException(status_code=403, detail='Payslip access denied')

    row = _get_scoped_payslip(db, current_user, payslip_id)

    return Response(
        content=row["file_content"],
        media_type=row.get('content_type') or 'application/zip',
        headers={
            "Content-Disposition": f'attachment; filename="{row["zip_filename"]}"'
        }
    )




@router.delete('/payslips/{payslip_id}')
def delete_payslip(payslip_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_tables(db)
    if not _can_payroll(db, current_user):
        raise HTTPException(status_code=403, detail='Only SuperUser, FranchiseUser or Finance users can delete payslips')
    _get_manageable_payslip(db, current_user, payslip_id)
    db.execute(text("""
        UPDATE payroll_payslips
        SET is_active = FALSE, deleted_at = :deleted_at, deleted_by_user_id = :deleted_by
        WHERE id = :id
    """), {'id': payslip_id, 'deleted_at': datetime.utcnow(), 'deleted_by': current_user.id})
    db.commit()
    return {'message': 'Payslip deleted'}
